import first_requests
import pandas as pd
import  openpyxl as ox


if __name__ == '__main__':

    # Параметры 'name_folder' и 'url' нужно задать вручную

    dict_link = {
        'ОБЪЕКТИВЫ': ["https://onfotolife.com/ru/cameras?type=compact&camera=Sea"]
    }
    
    category = []
    k = 0
    wb=ox.Workbook()
    ws = wb.worksheets[0]
    for name_folder, urls in dict_link.items(): 
        for url in urls:
            print(url)
            link_list = first_requests.get_links(url)['Ссылки']
            name_list = first_requests.get_links(url)['Названия']
            category.append(name_folder)

            # response = requests_web.responses(link_list, name_list, name_folder)
  
            for count in link_list: 
                value = count
                value2 = name_list[k]
                value3 = name_folder
                ws.cell(row=k+1, column=1).value = value
                ws.cell(row=k+1, column=2).value = value2
                ws.cell(row=k+1, column=3).value = value3
                k+=1 
                
    wb.save(r'C:\Users\Olga\Desktop\onfotolife\test.xlsx')

            # async_download.main_main(response[0], response[1])
