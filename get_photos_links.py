import time
from bs4 import BeautifulSoup
import asyncio
import aiohttp
import lxml
import  openpyxl as ox
from pprint import pprint


list_of_links_to_photo = []
list_of_names_photo = []

async def get_page_data(session, page, filename):
    
    # print(page)
    async with session.get(url=page) as response:
        responce_text = await response.text()
        soup = BeautifulSoup(responce_text, 'lxml')
        time.sleep(0.1)
        link = soup.find('div', class_='text-center').find('img')['src']
        list_of_links_to_photo.append(link)
        list_of_names_photo.append(filename)
        # print(filename)

    wb=ox.Workbook()
    ws = wb.worksheets[0]
    k = 0
    for count in list_of_links_to_photo: 
        value = count
        value2 = list_of_names_photo[k]
        ws.cell(row=k+1, column=1).value = value
        ws.cell(row=k+1, column=2).value = value2
        k+=1 
                
    wb.save(r'C:\Users\Olga\Desktop\onfotolife\финальные_ссылки.xlsx') 



async def gather_data(list_img_link, full_path_img_list):

    urls = list_img_link
    
    async with aiohttp.ClientSession() as session:

        tasks = []
        count = 0
        for page in urls:
            filename = full_path_img_list[count]
            task = asyncio.create_task(get_page_data(session, page, filename))
            count += 1
            tasks.append(task)
            time.sleep(0.1)

        await asyncio.gather(*tasks)
        
        

def main(list_img_link, full_path_img_list):
    asyncio.run(gather_data(list_img_link, full_path_img_list))


if __name__ == '__main__':
    
    workbook=ox.load_workbook(r'C:\Users\Olga\Desktop\onfotolife\список ссылок на страницы с фотографией.xlsx')
    worksheet = workbook.worksheets[0]
    list_img_link = []
    full_path_img_list = []
    for i in range(1, worksheet.max_row):
        list_img_link.append(worksheet.cell(row = i, column=2).value)
        full_path_img_list.append(worksheet.cell(row = i, column=1).value)
    main(list_img_link, full_path_img_list)
    workbook.close


    