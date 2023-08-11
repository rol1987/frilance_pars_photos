import asyncio
import httpx
import tqdm
import  openpyxl as ox
import time
import aiofiles
import requests
from concurrent.futures import ThreadPoolExecutor
import uuid



# async def download_files(url, filename):
#     # print('Запуск сохранения файла', filename, url)
#     with open(filename, 'wb') as f:
#         async with httpx.AsyncClient() as client:
#             async with client.stream('GET', url) as r:
#                 r.raise_for_status()
#                 time.sleep(0.2)
#                 for chuk in r.iter_bytes():
#                     f.write(chuk)

                
# async def main(urls, filename):

#     loop = asyncio.get_running_loop()
# #     task = [loop.create_task(download_files(url, filename)) for url, filename in urls]
# #     await asyncio.gather(*task, return_exceptions=True)
#     k=0
#     for url in urls:
#         tasks = loop.create_task(download_files(url, filename[k]))
#         k+=1

#     await asyncio.gather(*tasks, return_exceptions=True)



if __name__ == '__main__':
    
    workbook=ox.load_workbook(r'C:\Users\Olga\Desktop\onfotolife\финальные_ссылки.xlsx')
    worksheet = workbook.worksheets[0]
    links = []
    patchs = []
    for i in range(1, worksheet.max_row):
        links.append(worksheet.cell(row = i, column=1).value)
        patchs.append(worksheet.cell(row = i, column=2).value)
    # main(links, patchs)
    workbook.close

    # asyncio.run(main(links, patchs))

    urls = links

    def download(url):
        response = requests.get(url)

        # path_to_file = path
        path_to_file = f'c:\\Users\\Olga\\Desktop\\onfotolife\\temp\\{uuid.uuid1()}.jpg'

        with open(path_to_file, 'wb') as handle:
            handle.write(response.content)

        print(f'Файл {path_to_file} успешно скачан!')

    with ThreadPoolExecutor(max_workers=16) as executor:
        executor.map(download, urls)
    