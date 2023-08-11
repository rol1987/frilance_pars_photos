import grequests
import requests
import bs4
import lxml
import os
from pprint import pprint
from io import BytesIO
from PIL import Image
import time
import  openpyxl as ox


def responses(links, name_list, name_folder):
    # Запускаем цикл и проходим по первичным ссылкам
    number_link = 0 # нужно для того, чтобы итерироваться по списку и названиями страниц. Чтобы потом добавлять это в название сохраняемого файла
    list_img_link = []
    full_path_img_list = []
    sites = links
    response = (grequests.get(url) for url in sites)
    resp = grequests.map(response)
    for r in resp:    
        # Проверяем есть ли папка. Если нет - создаем её
        cwd = os.path.dirname(__file__)
        full_path = os.path.join(cwd, name_folder[number_link])
        full_path_temp = os.path.join(cwd, 'temp')  

        if not os.path.exists(full_path):
            os.makedirs(full_path)

        if not os.path.exists(full_path_temp):
            os.makedirs(full_path_temp)

        ind = 1
        number_pages = 1
        next_page = r.url
        pages_list = []

        pages_list.append(next_page)
        while ind == 1:
            
            next_page = next_page.replace(f'page={number_pages}', f'page={number_pages + 1}')
            
            # print('Пробуем подключиться к странице: ', next_page)

            r1 = requests.get(next_page)
            if r1.status_code == 200:
                soup = bs4.BeautifulSoup(r1.text, 'lxml')
                
                find_next_page = soup.find('a', {'href': {next_page}})
                if find_next_page != None:
                    if int(find_next_page.text) == number_pages + 1:
                        pages_list.append(next_page) # Список ссылок на внутренние страницы
                        number_pages += 1
                        ind = 1
                    else:
                        ind = 0
                else:
                    ind = 0
            else:
                ind = 0
    
        response = (grequests.get(url) for url in pages_list)
        resp_pages_list = grequests.map(response)
        

        num=0
        for rpl in resp_pages_list:
            
            soup_rpl = bs4.BeautifulSoup(rpl.text, 'lxml')            
            
            count_photo = 0
            for l in soup_rpl.find('div', class_='clearfix mosaicflow').find_all('a', class_='fancyimage'):
                count_photo += 1
                list_img_link.append(l['href']) # Список, в котором хранятся ссылки на все фотографии подраздела
                file_name = name_list[number_pages]
                file_name = file_name.replace('/', '_')
                full_path_img_list.append(os.path.join(full_path_temp, f'{file_name}_{num}.jpg'))
                print(os.path.join(full_path_temp, f'{file_name}_{num}.jpg'))
                num += 1

        number_link += 1 
    return full_path_img_list, list_img_link


if __name__ == '__main__':
    k = 0
    workbook=ox.load_workbook(r'C:\Users\Olga\Desktop\onfotolife\test.xlsx')
    worksheet = workbook.worksheets[0]
    
    links = []
    name_list = []
    name_folder = []
    for i in range(1, worksheet.max_row):
        links.append(worksheet.cell(row = i, column=1).value)
        name_list.append(worksheet.cell(row = i, column=2).value)
        name_folder.append(worksheet.cell(row = i, column=3).value)
    
    result = responses(links, name_list, name_folder)

    wb=ox.Workbook()
    ws = wb.worksheets[0]
    k = 0
    for count in result[0]: 
        value = count
        value2 = result[1][k]
        ws.cell(row=k+1, column=1).value = value
        ws.cell(row=k+1, column=2).value = value2
        k+=1 
                
    wb.save(r'C:\Users\Olga\Desktop\onfotolife\список ссылок на страницы с фотографией.xlsx') 